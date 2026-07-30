"""
send_emails.py
Reads output/manifest.csv (produced by generate_certificates.py), groups certificates
by POC email, and sends ONE email per POC with all their cluster's certificate PDFs attached.

Usage:
    python send_emails.py
Safety:
    While config.yaml -> email.dry_run: true, every email is sent to
    email.dry_run_recipient instead of the real POC, so you can verify
    subject/body/attachments before the real send.
"""

import os
import re
import smtplib
import ssl
from email.message import EmailMessage

import pandas as pd
import yaml


def load_config(path="config.yaml"):
    with open(path, "r") as f:
        return yaml.safe_load(f)


def safe_filename(s):
    """Strip characters that are unsafe in filenames."""
    return re.sub(r"[^A-Za-z0-9_\-]+", "_", str(s)).strip("_")


def generate_event_zips(group_df=None, pdf_paths=None, active_type="participation"):
    """
    Groups certificates by event_name and packages them into in-memory ZIP files.
    If in results mode and an event has multiple winning positions, places PDFs in position subfolders inside the ZIP.
    Returns dict mapping zip_filename -> zip_bytes.
    """
    import io
    import zipfile

    def format_pos_zip(p):
        p_str = str(p).strip()
        if not p_str or pd.isna(p) or p_str.lower() in ["nan", "none"]:
            return ""
        p_lower = p_str.lower()
        if "1" in p_lower or "first" in p_lower:
            return "1st Place"
        if "2" in p_lower or "second" in p_lower:
            return "2nd Place"
        if "3" in p_lower or "third" in p_lower:
            return "3rd Place"
        return p_str

    is_results = str(active_type).lower() in ["results", "winner", "merit"]

    zips = {}
    if group_df is not None and "event_name" in group_df.columns:
        used_names = {}
        for raw_ev, ev_group in group_df.groupby("event_name", dropna=False):
            ev_str = str(raw_ev).strip() if raw_ev and not pd.isna(raw_ev) and str(raw_ev).strip() else "General_Event"
            base_fname = safe_filename(ev_str) or "Event_Certificates"

            if base_fname in used_names:
                used_names[base_fname] += 1
                zip_filename = f"{base_fname}_{used_names[base_fname]}.zip"
            else:
                used_names[base_fname] = 1
                zip_filename = f"{base_fname}.zip"

            has_positions = is_results and "position" in ev_group.columns
            if has_positions:
                formatted_positions = ev_group["position"].apply(format_pos_zip)
                unique_positions = [p for p in formatted_positions.unique() if p]
            else:
                unique_positions = []

            buf = io.BytesIO()
            with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
                arc_names = set()
                for _, row in ev_group.iterrows():
                    pdf_p = row.get("pdf_path")
                    if pdf_p and os.path.exists(pdf_p):
                        base_arc = os.path.basename(pdf_p)
                        if has_positions and len(unique_positions) > 1:
                            pos_val = format_pos_zip(row.get("position"))
                            pos_sub = safe_filename(pos_val) if pos_val else ""
                            arc_name = f"{pos_sub}/{base_arc}" if pos_sub else base_arc
                        else:
                            arc_name = base_arc

                        if arc_name in arc_names:
                            b_name, ext_name = os.path.splitext(arc_name)
                            idx = 2
                            while f"{b_name}_{idx}{ext_name}" in arc_names:
                                idx += 1
                            arc_name = f"{b_name}_{idx}{ext_name}"
                        arc_names.add(arc_name)
                        zf.write(pdf_p, arc_name)
            zips[zip_filename] = buf.getvalue()
    elif pdf_paths:
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            arc_names = set()
            for pdf_p in pdf_paths:
                if pdf_p and os.path.exists(pdf_p):
                    arc_name = os.path.basename(pdf_p)
                    if arc_name in arc_names:
                        base_arc, ext_arc = os.path.splitext(arc_name)
                        idx = 2
                        while f"{base_arc}_{idx}{ext_arc}" in arc_names:
                            idx += 1
                        arc_name = f"{base_arc}_{idx}{ext_arc}"
                    arc_names.add(arc_name)
                    zf.write(pdf_p, arc_name)
        zips["Certificates.zip"] = buf.getvalue()

    return zips


def generate_excel_bytes(student_names, pdf_paths, group_df=None, active_type="participation"):
    """
    Generates a beautifully styled, fully interactive Excel (.xlsx) file for the POC email contingent.
    Ensures gridlines are visible, column widths auto-fit content (never chopped),
    headers are styled in Christ Blue (#004C6D), and Excel auto-filters are enabled.
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    excel_cols = {}
    excel_cols["S.No."] = list(range(1, len(student_names) + 1))
    excel_cols["Student Name"] = student_names

    if group_df is not None:
        if "school" in group_df.columns and not group_df["school"].dropna().empty:
            excel_cols["School / Institution"] = group_df["school"].fillna("").tolist()

        if "event_name" in group_df.columns:
            excel_cols["Event Name"] = group_df["event_name"].fillna("").tolist()

        # Position column is ONLY for results/winner/merit certificates
        if "position" in group_df.columns and str(active_type).lower() in ["results", "winner", "merit"]:
            def format_pos(p):
                p_str = str(p).strip()
                if not p_str or pd.isna(p) or p_str.lower() in ["nan", "none"]:
                    return ""
                p_lower = p_str.lower()
                if "1" in p_lower or "first" in p_lower:
                    return "1st Place"
                if "2" in p_lower or "second" in p_lower:
                    return "2nd Place"
                if "3" in p_lower or "third" in p_lower:
                    return "3rd Place"
                return p_str

            excel_cols["Position"] = [format_pos(x) for x in group_df["position"].tolist()]

    cert_files = [os.path.basename(p) if p else "" for p in pdf_paths]
    if len(cert_files) < len(student_names):
        cert_files.extend([""] * (len(student_names) - len(cert_files)))
    elif len(cert_files) > len(student_names):
        cert_files = cert_files[:len(student_names)]
    excel_cols["Certificate File"] = cert_files

    headers = list(excel_cols.keys())
    num_rows = len(student_names)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Certificates Roster"

    # Explicitly enable gridlines in Excel / Gmail web previewer
    ws.views.sheetView[0].showGridLines = True

    # Styling definitions
    header_fill = PatternFill(start_color="004C6D", end_color="004C6D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=11, color="1E293B")
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # Header row
    ws.append(headers)
    ws.row_dimensions[1].height = 26
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for i in range(num_rows):
        row_vals = [excel_cols[k][i] for k in headers]
        ws.append(row_vals)
        row_idx = i + 2
        ws.row_dimensions[row_idx].height = 20
        for col_idx, col_name in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            align = "center" if col_name in ["S.No.", "Position"] else "left"
            cell.alignment = Alignment(horizontal=align, vertical="center")

    # Dynamic Column Width Auto-Fitting
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    # Enable native Excel Auto-Filter
    if num_rows > 0:
        ws.auto_filter.ref = ws.dimensions

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    return excel_buffer.getvalue()


def build_email(poc_name, poc_email, cluster_id, event_name, student_names, pdf_paths, cfg, group_df=None, drive_data=None):
    at = str(cfg.get("active_type", "participation")).lower()
    active_cfg = cfg.get("types", {}).get(at, {})
    global_ecfg = cfg.get("email", {})
    type_ecfg = active_cfg.get("email", {})
    ecfg = {**global_ecfg, **type_ecfg}

    is_direct = at in ["school", "volunteer"]
    poc_folder_url = (drive_data.get("poc_folder_url") if drive_data else None) if not is_direct else None

    # Event summary (for bulk types)
    if not is_direct and group_df is not None and "event_name" in group_df.columns and not group_df["event_name"].dropna().empty:
        text_lines = ["Events Summary:"]
        html_items = []
        for raw_ev, ev_group in group_df.groupby("event_name", dropna=False):
            ev_display = str(raw_ev).strip() if raw_ev and not pd.isna(raw_ev) and str(raw_ev).strip() else "General Event"
            count = len(ev_group)
            text_lines.append(f"  • {ev_display} ({count} certificate{'s' if count != 1 else ''})")
            html_items.append(
                f"<li style='margin-bottom: 6px; font-weight: 600; color: #1e293b;'>{ev_display} <span style='font-weight: 400; color: #64748b;'>({count} certificate{'s' if count != 1 else ''})</span></li>"
            )
        student_list_text = "\n".join(text_lines)
        html_student_list = (
            "<div style='margin: 18px 0; background-color: #f8fafc; border: 1px solid #e2e8f0; border-radius: 6px; padding: 14px 18px; font-family: -apple-system, BlinkMacSystemFont, sans-serif;'>"
            "<div style='font-size: 12px; font-weight: 700; color: #004C6D; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 8px;'>Events Summary</div>"
            "<ul style='margin: 0; padding-left: 18px; font-size: 14px; font-family: -apple-system, BlinkMacSystemFont, sans-serif;'>"
            + "".join(html_items)
            + "</ul></div>"
        )
    elif not is_direct:
        total_certs = len(student_names)
        student_list_text = f"Total Certificates: {total_certs}"
        html_student_list = (
            f"<div style='margin: 16px 0; font-family: -apple-system, BlinkMacSystemFont, sans-serif; font-size: 14px; color: #334155; font-weight: 600;'>"
            f"Total Certificates Enclosed: {total_certs}"
            f"</div>"
        )
    else:
        student_list_text = ""
        html_student_list = ""

    # Google Drive CTA for bulk types
    drive_card_html = ""
    if poc_folder_url and not is_direct:
        drive_card_html = f"""
        <div style="background-color: #f8fafc; border: 1px solid #cbd5e1; border-radius: 8px; padding: 22px 20px; margin: 24px 0; text-align: center; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Helvetica, Arial, sans-serif;">
            <div style="font-size: 16px; font-weight: 700; color: #0f172a; margin-bottom: 6px;">
                Google Drive Certificate Repository
            </div>
            <div style="font-size: 13px; color: #475569; margin-bottom: 18px; line-height: 1.5;">
                All certificate PDFs have been structured by event in your Google Drive directory.
            </div>
            <a href="{poc_folder_url}" target="_blank" style="display: inline-block; background-color: #004C6D; color: #ffffff; text-decoration: none; padding: 11px 26px; font-weight: 600; border-radius: 6px; font-size: 14px; letter-spacing: 0.2px;">
                View All Certificates on Google Drive
            </a>
        </div>
        """

    school_val = (
        group_df["school"].iloc[0]
        if (group_df is not None and "school" in group_df.columns and not group_df["school"].dropna().empty)
        else (poc_name or poc_email)
    )

    fmt_vars_text = {
        "poc_name": poc_name or poc_email,
        "poc_email": poc_email,
        "volunteer_name": poc_name or poc_email,
        "volunteer_email": poc_email,
        "school": school_val,
        "cluster_id": cluster_id or "",
        "event_name": event_name or "",
        "student_list": student_list_text,
    }

    try:
        subject = ecfg["subject_template"].format_map(fmt_vars_text)
    except Exception:
        subject = f"ANVESHA '26 | Participation Certificates & Contingent Summary - {school_val}"

    raw_body_tmpl = ecfg.get("body_template", "")
    if not is_direct and "{student_list}" not in raw_body_tmpl:
        raw_body_tmpl = raw_body_tmpl.strip() + "\n\n{student_list}"

    try:
        body_text = raw_body_tmpl.format_map(fmt_vars_text)
    except Exception:
        body_text = f"Dear {poc_name or poc_email},\n\nPlease find your certificate attached."

    if poc_folder_url and not is_direct:
        body_text += f"\n\nGoogle Drive Folder Link:\n{poc_folder_url}\n"

    fmt_vars_html = {
        "poc_name": f"<strong>{poc_name or poc_email}</strong>",
        "poc_email": poc_email,
        "volunteer_name": f"<strong>{poc_name or poc_email}</strong>",
        "volunteer_email": poc_email,
        "school": f"<strong>{school_val}</strong>",
        "cluster_id": cluster_id or "",
        "event_name": f"<strong>{event_name or ''}</strong>",
        "student_list": html_student_list,
    }

    try:
        body_html_content = raw_body_tmpl.format_map(fmt_vars_html)
    except Exception:
        body_html_content = f"Dear <strong>{poc_name or poc_email}</strong>,<br><br>Please find your certificate attached."

    # Locate Thank You graphic if present
    ty_img_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "htmlbodymail", "Anvesha Participant Appreciation.jpg.jpeg")
    if not os.path.exists(ty_img_path):
        ty_img_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sample_data", "Anvesha Participant Appreciation.jpg.jpeg")
    has_ty_img = os.path.exists(ty_img_path)

    thank_you_banner_html = (
        '<div style="text-align: center; margin: 28px 0 24px 0;">'
        '<img src="cid:thank_you_banner" alt="Thank You - ANVESHA \'26" '
        'style="width: 100%; max-width: 590px; height: auto; border-radius: 10px; '
        'box-shadow: 0 4px 14px rgba(0,0,0,0.12); border: 1px solid #e2e8f0; display: block; margin: 0 auto;" />'
        '</div>'
    ) if has_ty_img else ""

    # Format the paragraphs for clean HTML
    paragraphs = [p.strip() for p in body_html_content.split("\n\n") if p.strip()]
    formatted_paragraphs = []
    ty_inserted = False
    for p in paragraphs:
        if html_student_list and html_student_list in p:
            formatted_paragraphs.append(drive_card_html + p)
            if thank_you_banner_html:
                formatted_paragraphs.append(thank_you_banner_html)
                ty_inserted = True
        else:
            p_clean = p.replace("\n", "<br>")
            p_clean = p_clean.replace("ANVESHA '26", "<strong>ANVESHA '26</strong>")
            p_clean = p_clean.replace("ANVESHA ''26", "<strong>ANVESHA '26</strong>")
            p_clean = p_clean.replace("CHRIST (Deemed to be University), Kengeri Campus", "<strong>CHRIST (Deemed to be University), Kengeri Campus</strong>")
            formatted_paragraphs.append(f"<p style='margin: 0 0 16px 0; font-family: Georgia, serif; font-size: 15px; line-height: 1.6;'>{p_clean}</p>")

    if drive_card_html and not any(drive_card_html in fp for fp in formatted_paragraphs):
        formatted_paragraphs.insert(0, drive_card_html)

    if thank_you_banner_html and not ty_inserted:
        if len(formatted_paragraphs) > 1:
            formatted_paragraphs.insert(-1, thank_you_banner_html)
        else:
            formatted_paragraphs.append(thank_you_banner_html)

    html_body_paragraphs = "".join(formatted_paragraphs)

    html_email = f"""<!DOCTYPE html>
<html>
<head>
  <meta charset="utf-8">
  <title>{subject}</title>
</head>
<body style="margin: 0; padding: 0; background-color: #F2F2F2; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Helvetica, Arial, sans-serif;">
  <table bgcolor="#2E2E2E" border="0" cellpadding="0" cellspacing="0" style="background-image: url('https://kp.christuniversity.in/KnowledgePro/BulkEmail/images/baby-blue-.jpg'); background-size: cover; background-position: center; background-attachment: fixed; background-color: #F2F2F2; padding: 40px 10px;" width="100%"> 
   <tbody> 
    <tr> 
     <td align="center" valign="top" width="100%"> 
      <table border="0" cellpadding="0" cellspacing="0" width="100%" style="max-width: 650px; background-color: rgba(0, 0, 0, 0.5); border-radius: 8px; overflow: hidden;"> 
       <tbody> 
        <tr> 
         <td align="center" class="header" style="color: #4C4C4C; background-color: #004C6D;" valign="top">
           <img alt="CHRIST Deemed to be University" src="cid:christ_header" style="width: 100%; max-width: 650px; height: auto; display: block; border: 0;">
         </td> 
        </tr> 
        <tr> 
         <td align="center" class="content" style="border: 0px none transparent; padding: 20px 20px;" valign="top"> 
          <table border="0" cellpadding="0" cellspacing="0" width="100%"> 
           <tbody> 
            <tr> 
             <td align="left" class="mainContent" style="background-color: #ffffff; padding: 30px 25px; border-radius: 8px; box-shadow: 0 2px 10px rgba(0, 0, 0, 0.1); font-family: 'Georgia', serif; font-size: 15px; color: #333333; line-height: 1.6;" valign="top"> 
              <div style="text-align: center; font-family: 'Georgia', serif; font-size: 22px; color: #004C6D; margin-bottom: 20px; font-weight: bold; border-bottom: 1px solid #e2e8f0; padding-bottom: 12px;">
               ANVESHA 2026
              </div> 
              {html_body_paragraphs}
             </td> 
            </tr> 
           </tbody> 
          </table> </td> 
        </tr> 
        <tr> 
         <td align="center" class="socialLinks" style="padding: 20px 10px; background-color: #004C6D;" valign="top"> 
           <p style="margin: 0 0 10px 0;">
             <a href="https://facebook.com/christuniversityblr" target="_blank"><img src="https://kp.christuniversity.in/KnowledgePro/images/BulkEmail/images/1784293302306fb.png" style="width: 35px; height: 35px; border: none; padding: 0 8px;"></a> 
             <a href="https://www.instagram.com/christuniversity_kengeri/" target="_blank"><img src="https://kp.christuniversity.in/KnowledgePro/images/BulkEmail/images/1784293302306Insta.png" style="width: 35px; height: 35px; border: none; padding: 0 8px;"></a> 
             <a href="https://www.youtube.com/@christuniversitykengeri" target="_blank"><img src="https://kp.christuniversity.in/KnowledgePro/images/BulkEmail/images/1784293302306youtube.png" style="width: 35px; height: 35px; border: none; padding: 0 8px;"></a> 
             <a href="https://linkedin.com/company/christ-university-faculty-of-engineering" target="_blank"><img src="https://kp.christuniversity.in/KnowledgePro/images/BulkEmail/images/1784293302306linkedin%20(2).png" style="width: 35px; height: 35px; border: none; padding: 0 8px;"></a> 
             <a href="https://www.flickr.com/photos/199433161@N02/albums" target="_blank"><img src="https://kp.christuniversity.in/KnowledgePro/images/BulkEmail/images/1784293302306flicker.png" style="width: 35px; height: 35px; border: none; padding: 0 8px;"></a>
           </p> 
           <p style="margin: 0; font-family: Arial, sans-serif; font-size: 13px;"><a href="http://www.christuniversity.in" target="_blank" style="text-decoration: none;"><span style="color:#FFFFFF; font-weight: bold;">www.christuniversity.in</span></a></p> 
         </td> 
        </tr> 
       </tbody> 
      </table> </td> 
    </tr> 
   </tbody> 
  </table> 
 </body>
</html>"""

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = f"{ecfg['sender_name']} <{ecfg['sender_email']}>"

    dry_run = ecfg.get("dry_run", False)
    dry_run_recipient = ecfg.get("dry_run_recipient", "")
    msg["To"] = dry_run_recipient if (dry_run and dry_run_recipient) else poc_email
    
    msg.set_content(body_text)
    msg.add_alternative(html_email, subtype="html")

    # Add inline header image
    header_img_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sample_data", "christ_header.png")
    if os.path.exists(header_img_path):
        with open(header_img_path, "rb") as f:
            img_data = f.read()
        html_part = None
        for part in msg.iter_parts():
            if part.get_content_type() == "multipart/alternative":
                for subpart in part.iter_parts():
                    if subpart.get_content_type() == "text/html":
                        html_part = subpart
                        break
            elif part.get_content_type() == "text/html":
                html_part = part
                break
        
        if html_part:
            html_part.add_related(
                img_data,
                maintype="image",
                subtype="png",
                cid="christ_header"
            )
        else:
            msg.add_attachment(
                img_data,
                maintype="image",
                subtype="png",
                cid="christ_header",
                filename="christ_header.png"
            )

    # Add inline Thank You appreciation image if present
    if has_ty_img:
        with open(ty_img_path, "rb") as f:
            ty_data = f.read()
        html_part = None
        for part in msg.iter_parts():
            if part.get_content_type() == "multipart/alternative":
                for subpart in part.iter_parts():
                    if subpart.get_content_type() == "text/html":
                        html_part = subpart
                        break
            elif part.get_content_type() == "text/html":
                html_part = part
                break
        
        if html_part:
            html_part.add_related(
                ty_data,
                maintype="image",
                subtype="jpeg",
                cid="thank_you_banner"
            )
        else:
            msg.add_attachment(
                ty_data,
                maintype="image",
                subtype="jpeg",
                cid="thank_you_banner",
                filename="thank_you_banner.jpeg"
            )

    if is_direct:
        # Direct delivery: attach PDF certificate directly
        for pdf_p in pdf_paths:
            if pdf_p and os.path.exists(pdf_p):
                with open(pdf_p, "rb") as f:
                    pdf_bytes = f.read()
                msg.add_attachment(
                    pdf_bytes,
                    maintype="application",
                    subtype="pdf",
                    filename=os.path.basename(pdf_p)
                )
    else:
        # Bulk delivery: attach Excel sheet + ZIP files
        try:
            excel_bytes = generate_excel_bytes(student_names, pdf_paths, group_df, active_type=at)
            if excel_bytes:
                msg.add_attachment(
                    excel_bytes,
                    maintype="application",
                    subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    filename="student_list.xlsx"
                )
        except Exception as exc:
            print(f"Warning: Could not attach student_list.xlsx: {exc}")

        # Attach event-grouped ZIP files ONLY IF Drive upload is not used
        if not poc_folder_url:
            event_zips = generate_event_zips(group_df, pdf_paths, active_type=at)
            total_zip_bytes = sum(len(b) for b in event_zips.values())
            if total_zip_bytes > 18 * 1024 * 1024:
                mb_size = round(total_zip_bytes / (1024 * 1024), 1)
                raise ValueError(
                    f"Total attachment size ({mb_size} MB) exceeds Gmail's 25 MB limit. "
                    "Google Drive upload was disabled or failed. Please check Google Drive setup or send via fallback folder."
                )
            for zip_filename, zip_bytes in event_zips.items():
                msg.add_attachment(
                    zip_bytes,
                    maintype="application",
                    subtype="zip",
                    filename=zip_filename
                )

    return msg


def main():
    cfg = load_config()
    manifest_path = os.path.join(cfg["output_dir"], "manifest.csv")

    if not os.path.exists(manifest_path):
        print(f"No manifest found at {manifest_path}. Run generate_certificates.py first.")
        return

    df = pd.read_csv(manifest_path)
    df = df[df["pdf_path"].apply(os.path.exists)]  # only send certs that actually exist on disk

    ecfg = cfg["email"]
    if ecfg["dry_run"]:
        print(f"*** DRY RUN MODE: all emails will go to {ecfg['dry_run_recipient']} instead of real POCs ***")
        print("*** Set email.dry_run: false in config.yaml once you've verified output. ***\n")

    context = ssl.create_default_context()
    sent, failed = 0, 0
    port = int(ecfg.get("smtp_port", 465))

    def _make_connection():
        if port == 465:
            srv = smtplib.SMTP_SSL(ecfg["smtp_host"], port, timeout=60, context=context)
        else:
            srv = smtplib.SMTP(ecfg["smtp_host"], port, timeout=60)
            srv.ehlo()
            srv.starttls(context=context)
            srv.ehlo()
        srv.login(ecfg["sender_email"], ecfg["sender_app_password"])
        return srv

    import time
    for poc_email, group in df.groupby("poc_email"):
        poc_name = group["poc_name"].iloc[0] if "poc_name" in group.columns else group["poc_email"].iloc[0]
        cluster_id = group["cluster_id"].iloc[0] if "cluster_id" in group.columns else ""
        event_name = group["event_name"].iloc[0] if "event_name" in group.columns else ""
        student_names = group["student_name"].tolist() if "student_name" in group.columns else group["school"].tolist() if "school" in group.columns else ["Certificate"] * len(group)
        pdf_paths = group["pdf_path"].tolist()

        msg = build_email(poc_name, poc_email, cluster_id, event_name, student_names, pdf_paths, cfg, group)

        try:
            srv = _make_connection()
            try:
                srv.send_message(msg)
                print(f"Sent -> {msg['To']}  ({len(pdf_paths)} certificate(s) for {cluster_id})")
                sent += 1
            finally:
                try:
                    srv.quit()
                except Exception:
                    pass
        except Exception as e:
            print(f"FAILED -> {poc_email}: {e}")
            failed += 1
            
            # ── Fallback: save PDFs + message.html + christ_header.png + student_list.xlsx so user can send manually ──
            try:
                import shutil
                from pathlib import Path
                ROOT_DIR = Path(__file__).parent
                fallback_dir = ROOT_DIR / "fallback" / str(poc_email)
                fallback_dir.mkdir(parents=True, exist_ok=True)

                # Save event-grouped ZIP files into the fallback folder
                event_zips = generate_event_zips(group, pdf_paths)
                for zip_name, zip_bytes in event_zips.items():
                    (fallback_dir / zip_name).write_bytes(zip_bytes)

                # Copy every PDF into the fallback folder
                for pdf_path in pdf_paths:
                    if os.path.exists(pdf_path):
                        shutil.copy2(pdf_path, fallback_dir / os.path.basename(pdf_path))

                # Copy the logo header banner for offline rendering
                header_src = ROOT_DIR / "sample_data" / "christ_header.png"
                if header_src.exists():
                    shutil.copy2(header_src, fallback_dir / "christ_header.png")

                # Save the student list Excel sheet in fallback folder
                try:
                    excel_bytes = generate_excel_bytes(student_names, pdf_paths, group, active_type=cfg.get("active_type", "participation"))
                    if excel_bytes:

                        (fallback_dir / "student_list.xlsx").write_bytes(excel_bytes)
                except Exception as fb_xl_exc:
                    print(f"  -> WARNING: Could not save fallback Excel sheet: {fb_xl_exc}")

                subject = str(msg["Subject"]) if msg["Subject"] else "(no subject)"
                
                # 1. Plain text fallback message
                body_part = msg.get_body(preferencelist=("plain",))
                body_text = body_part.get_content() if body_part else "(no body)"
                message_txt = (
                    f"TO: {poc_email}\n"
                    f"SUBJECT: {subject}\n"
                    f"{'=' * 60}\n\n"
                    f"{body_text.strip()}\n\n"
                    f"{'=' * 60}\n"
                    f"ATTACHMENTS ({len(pdf_paths) + 1}):\n"
                    f"  - student_list.xlsx\n"
                    + "\n".join(f"  - {os.path.basename(p)}" for p in pdf_paths)
                )
                (fallback_dir / "message.txt").write_text(message_txt, encoding="utf-8")

                # 2. HTML fallback message (perfect for copying/pasting directly into Gmail)
                html_part = msg.get_body(preferencelist=("html",))
                if html_part:
                    html_content = html_part.get_content()
                    # Convert header image to Base64 data URL so it renders properly when copy-pasted to email clients
                    header_src = ROOT_DIR / "sample_data" / "christ_header.png"
                    if header_src.exists():
                        import base64
                        try:
                            b64_data = base64.b64encode(header_src.read_bytes()).decode("utf-8")
                            html_for_copy = html_content.replace("cid:christ_header", f"data:image/png;base64,{b64_data}")
                        except Exception:
                            html_for_copy = html_content.replace("cid:christ_header", "christ_header.png")
                    else:
                        html_for_copy = html_content.replace("cid:christ_header", "christ_header.png")
                    (fallback_dir / "message.html").write_text(html_for_copy, encoding="utf-8")

                print(f"  -> Fallback saved: fallback/{poc_email}/ ({len(pdf_paths)} PDF(s) + student_list.xlsx + message.html)")
            except Exception as fb_exc:
                print(f"  -> WARNING: Could not write fallback: {fb_exc}")

        time.sleep(1)  # stay within Gmail rate limits

    print(f"\nDone. {sent} email(s) sent, {failed} failed.")


if __name__ == "__main__":
    main()
